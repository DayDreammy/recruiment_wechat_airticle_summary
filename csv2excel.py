import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink

# 定义信源文件地址
source_file_path = r'D:\wechatmonitor\source\招聘公众号信源.txt'


def adjust_cell_format(cell, r_idx, c_idx, link_column_index, summary_column_index):
    # 如果是超链接列并且不是第一行，则设置单元格为超链接
    if c_idx == link_column_index and r_idx != 1:
        cell.hyperlink = Hyperlink(
            ref=cell.coordinate, target=cell.value, tooltip=cell.value)
        cell.style = 'Hyperlink'  # 应用超链接的默认格式
    # 如果不是第一行，则设置单元格格式为左右上下居中对齐
    if r_idx != 1:
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrapText=True)
    # 如果是摘要列并且不是第一行，则设置单元格为左对齐
    if c_idx == summary_column_index and r_idx != 1:
        cell.alignment = Alignment(
            horizontal='left', vertical='center', wrapText=True)


def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列名
        for cell in col:
            try:  # 避免空单元格引发错误
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        # 设置最大列宽为180
        if adjusted_width > 180:
            adjusted_width = 180
        ws.column_dimensions[column].width = adjusted_width


def adjust_row_height(ws):
    for row in ws.rows:
        max_height = 20
        for cell in row:
            try:  # 避免空单元格引发错误
                if len(str(cell.value)) > max_height:
                    max_height = len(cell.value)
            except:
                pass
        adjusted_height = (max_height + 2) * 1.2
        ws.row_dimensions[row[0].row].height = adjusted_height


def sort_dataframe(df, sort_column_index):
    # 读取排序依据的文本文件
    with open(source_file_path, 'r', encoding='utf-8') as file:
        sort_list = file.read().splitlines()

    # 创建排序映射字典，未在列表中的元素将放在最后并按原顺序排序
    sort_order = {key: i for i, key in enumerate(sort_list)}
    max_index = len(sort_order)

    # 根据文本文件中的顺序对DataFrame进行排序
    df['SortKey'] = df.iloc[:, sort_column_index].map(sort_order)
    df['SortKey'].fillna(max_index + len(df), inplace=True)
    df.sort_values(by='SortKey', inplace=True)
    df.drop('SortKey', axis=1, inplace=True)

    return df


def csv_to_excel(csv_file_path, excel_file_path):
    link_column_index = 5  # 超链接所在列的索引
    summary_column_index = 6  # 摘要所在列的索引
    sort_column_index = 1  # 未添加序号列前，根据第B列进行排序，Python中索引为1

    # 读取CSV文件
    df = pd.read_csv(csv_file_path, quotechar='"', quoting=1)

    # 按指定顺序对DataFrame进行排序
    df = sort_dataframe(df, sort_column_index)

    # 在DataFrame首列添加序号列，序号从1开始
    df.insert(0, '序号', list(range(1, len(df) + 1)))

    # 创建一个新的Workbook
    wb = Workbook()
    ws = wb.active

    # 将DataFrame转换为Excel行
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            adjust_cell_format(cell, r_idx, c_idx,
                               link_column_index, summary_column_index)

    # 设置第一行的格式：加粗，字号大一点
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    adjust_column_width(ws)
    adjust_row_height(ws)

    # 保存Workbook到Excel文件
    wb.save(excel_file_path)

    print(f'CSV文件已成功转换为Excel文件并保存到 {excel_file_path}')
